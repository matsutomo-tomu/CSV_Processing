import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

def collect_csv_data():
    """シート名とCSVデータを収集する関数"""
    sheet_data = {}
    
    while True:
        # シート名の入力
        sheet_name = input("作成するシート名を入力してください（終了するには 'end' と入力）: ").strip()
        if sheet_name.lower() == "end":
            break

        if not sheet_name:
            print("エラー: シート名が空です。再入力してください。")
            continue
        if sheet_name in sheet_data:
            print("エラー: 既に存在するシート名です。別の名前を指定してください。")
            continue
        
        combined_data = []
        print(f"--- {sheet_name} のデータを入力してください ---")

        while True:
            # CSVファイルのパス入力
            csv_file = input(f"{sheet_name}に記入するCSVファイルのパスを入力してください（終了するには 'end' と入力）: ").strip().strip('"')
            if csv_file.lower() == "end":
                break

            # ファイル存在チェック
            if not os.path.isfile(csv_file):
                print(f"エラー: CSVファイルが見つかりません -> {csv_file}")
                continue

            try:
                # CSV読み込み（ヘッダーなし）
                df = pd.read_csv(csv_file, header=None)
                combined_data.append(df)
                # 1行分の空白を挿入
                empty_row = pd.DataFrame([[""] * len(df.columns)], columns=df.columns)
                combined_data.append(empty_row)

                print(f"CSVファイルを追加しました -> {csv_file}")
            except Exception as e:
                print(f"エラー: ファイルの処理中に問題が発生しました -> {e}")

        if combined_data:
            sheet_data[sheet_name] = combined_data
    
    return sheet_data

# 保存するExcelファイル名
output_excel = input("出力するExcelファイルの名前を入力（拡張子は省略可）→→ ").strip().strip('"')
if not output_excel.endswith(".xlsx"):
    output_excel += ".xlsx"

# 保存先フォルダ指定
save_dir = input("Excelファイルを保存するフォルダのパス入力してください（未入力の場合は現在のフォルダ）→→ ").strip().strip('"')
if not save_dir:
    save_dir = os.getcwd()
elif not os.path.isdir(save_dir):
    print(f"エラー: 指定したフォルダが存在しません -> {save_dir}")
    exit()

output_path = os.path.join(save_dir, output_excel)

# シートごとのデータを収集
sheet_data = collect_csv_data()

# Excelファイル書き出し
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for sheet_name, data in sheet_data.items():
        sheet_df = pd.concat(data, ignore_index=True)
        sheet_df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)  # ヘッダーなし

# 罫線を削除する
wb = load_workbook(output_path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border()  # 罫線なしに設定
wb.save(output_path)

print(f"Excelファイルが作成されました: {output_path}")
